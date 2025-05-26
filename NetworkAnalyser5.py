import sys
import csv
import time
import threading
from collections import deque
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QTableWidget, QTableWidgetItem,
    QLabel, QLineEdit, QHeaderView, QStatusBar, QHBoxLayout, QPushButton, QGraphicsOpacityEffect
)
from PyQt5.QtCore import Qt, pyqtSignal, QObject, QTimer, QPropertyAnimation, QEasingCurve
from PyQt5.QtGui import QColor, QPixmap
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
from scapy.all import sniff, IP, TCP, UDP

SUSPICIOUS_IPS = {"192.168.1.66", "45.83.122.5", "10.0.0.99"}
connection_tracker = {}
EXCEL_FILE = "packet_log.xlsx"

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"
    ws.append(["Time", "Source IP", "Destination IP", "Protocol", "Port", "Flags"])
    wb.save(EXCEL_FILE)

# Thread-safe signals
class PacketSignal(QObject):
    new_packet = pyqtSignal(dict)
    alert_signal = pyqtSignal(str)

signal_manager = PacketSignal()

class RealTimePlot(FigureCanvas):
    def __init__(self, parent=None, max_points=60):
        plt.style.use('dark_background')
        self.fig = Figure(figsize=(8, 4), facecolor='#16213e')
        self.ax1 = self.fig.add_subplot(121, facecolor='#0f3460')
        self.ax2 = self.fig.add_subplot(122, facecolor='#0f3460')
        super().__init__(self.fig)

        self.max_points = max_points
        self.packet_counts = deque([0]*max_points, maxlen=max_points)
        self.timestamps = deque([i for i in range(-max_points+1, 1)], maxlen=max_points)

        self.filtered_protocol_counts = {"TCP": 0, "UDP": 0, "OTHER": 0}

        self.line_plot, = self.ax1.plot(self.timestamps, self.packet_counts, 'cyan')
        
        # Customize plot appearance
        self.ax1.tick_params(colors='white')
        self.ax2.tick_params(colors='white')
        
        for spine in self.ax1.spines.values():
            spine.set_color('white')
        for spine in self.ax2.spines.values():
            spine.set_color('white')
            
        self.ax1.title.set_color('white')
        self.ax2.title.set_color('white')
        self.ax1.xaxis.label.set_color('white')
        self.ax1.yaxis.label.set_color('white')

        self.ax1.set_title("Packets / Second", fontsize=9)
        self.ax1.set_xlabel("Seconds Ago")
        self.ax1.set_ylabel("Count")
        self.ax1.grid(True, color='#333333')

        self.ax2.set_title("Protocol Breakdown (Filtered)", fontsize=9)

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_plot)
        self.timer.start(1000)

    def increment(self):
        self.packet_counts[-1] += 1

    def update_protocol_counts(self, protocols):
        self.filtered_protocol_counts = {"TCP": 0, "UDP": 0, "OTHER": 0}
        for p in protocols:
            if p in self.filtered_protocol_counts:
                self.filtered_protocol_counts[p] += 1
            else:
                self.filtered_protocol_counts["OTHER"] += 1

    def update_plot(self):
        self.packet_counts.append(0)
        self.timestamps.append(self.timestamps[-1] + 1)
        self.line_plot.set_ydata(self.packet_counts)
        self.line_plot.set_xdata(self.timestamps)
        self.ax1.relim()
        self.ax1.autoscale_view()

        self.ax2.clear()
        self.ax2.set_title("Protocol Breakdown (Filtered)", fontsize=9)
        protocols = list(self.filtered_protocol_counts.keys())
        values = list(self.filtered_protocol_counts.values())
        self.ax2.bar(protocols, values, color=['orange', 'purple', 'gray'])

        self.draw()

class NetworkAnalyzerGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🔥 Advanced Network Analyzer")
        self.setGeometry(100, 100, 1500, 800)

        # Custom title bar
        self.setWindowFlags(Qt.FramelessWindowHint)
        
        # Create title bar widget
        title_bar = QWidget()
        title_bar.setStyleSheet("background-color: #0f3460;")
        title_bar.setFixedHeight(40)
        
        title_layout = QHBoxLayout(title_bar)
        title_layout.setContentsMargins(10, 0, 10, 0)
        
        # Add logo and title
        self.logo_label = QLabel()
        # Create a simple placeholder logo if you don't have one
        pixmap = QPixmap(30, 30)
        pixmap.fill(QColor(240, 84, 84))  # Red square as placeholder
        self.logo_label.setPixmap(pixmap)
        
        self.title_label = QLabel("Advanced Network Analyzer")
        self.title_label.setStyleSheet("color: white; font-size: 16px; font-weight: bold;")
        
        # Add minimize/close buttons
        self.minimize_btn = QPushButton("─")
        self.maximize_btn = QPushButton("🗖")
        self.close_btn = QPushButton("✕")
        
        for btn in [self.minimize_btn, self.close_btn, self.maximize_btn]:
            btn.setFixedSize(30, 30)
            btn.setStyleSheet("""
                QPushButton {
                    color: white;
                    border: none;
                    font-size: 16px;
                }
                QPushButton:hover {
                    background-color: rgba(255, 255, 255, 0.2);
                }
            """)
        
        self.minimize_btn.clicked.connect(self.showMinimized)
        self.maximize_btn.clicked.connect(self.showMaximized)
        self.close_btn.clicked.connect(self.close)
        
        title_layout.addWidget(self.logo_label)
        title_layout.addWidget(self.title_label)
        title_layout.addStretch()
        title_layout.addWidget(self.minimize_btn)
        title_layout.addWidget(self.maximize_btn)
        title_layout.addWidget(self.close_btn)

        # Main content widgets
        self.packet_table = QTableWidget()
        self.packet_table.setColumnCount(6)
        self.packet_table.setHorizontalHeaderLabels(["Source IP", "Destination IP", "Protocol", "Port", "Flags", "Time"])
        self.packet_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.packet_table.setEditTriggers(QTableWidget.NoEditTriggers)

        self.alert_label = QLabel("")
        self.alert_label.setStyleSheet("""
            color: #f05454;
            font-weight: bold;
            font-size: 16px;
            background-color: rgba(240, 84, 84, 0.1);
            padding: 8px;
            border-left: 4px solid #f05454;
            border-radius: 2px;
        """)

        self.filter_bar = QLineEdit()
        self.filter_bar.setPlaceholderText("🔍 Filter by IP address...")
        self.filter_bar.textChanged.connect(self.filter_table)

        self.traffic_plot = RealTimePlot()

        # Create sidebar
        self.sidebar = QWidget()
        self.sidebar.setFixedWidth(200)
        self.sidebar.setStyleSheet("background-color: #0f3460;")
        
        sidebar_layout = QVBoxLayout()
        sidebar_layout.setContentsMargins(10, 20, 10, 20)
        
        # Add widgets to sidebar
        self.stats_title = QLabel("NETWORK STATS")
        self.stats_title.setStyleSheet("color: white; font-size: 14px; font-weight: bold;")
        self.stats_title.setAlignment(Qt.AlignCenter)
        
        self.top_sources = QLabel("Top Sources:\n- Loading...")
        self.top_destinations = QLabel("Top Destinations:\n- Loading...")
        self.protocol_dist = QLabel("Protocol Distribution:\n- Loading...")
        
        for label in [self.top_sources, self.top_destinations, self.protocol_dist]:
            label.setStyleSheet("color: white; font-size: 12px;")
            label.setWordWrap(True)
        
        sidebar_layout.addWidget(self.stats_title)
        sidebar_layout.addWidget(self.top_sources)
        sidebar_layout.addWidget(self.top_destinations)
        sidebar_layout.addWidget(self.protocol_dist)
        sidebar_layout.addStretch()
        
        self.sidebar.setLayout(sidebar_layout)

        # Add status bar
        self.status_bar = self.statusBar()
        self.packet_count_label = QLabel("Packets: 0")
        self.threat_count_label = QLabel("Threats: 0")
        self.connection_label = QLabel("Active Connections: 0")
        
        self.status_bar.addPermanentWidget(self.packet_count_label)
        self.status_bar.addPermanentWidget(self.threat_count_label)
        self.status_bar.addPermanentWidget(self.connection_label)

        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        main_layout.addWidget(title_bar)
        
        content_layout = QHBoxLayout()
        content_layout.addWidget(self.sidebar)
        
        right_layout = QVBoxLayout()
        right_layout.addWidget(self.filter_bar)
        right_layout.addWidget(self.packet_table)
        right_layout.addWidget(self.alert_label)
        right_layout.addWidget(self.traffic_plot)
        
        content_layout.addLayout(right_layout)
        main_layout.addLayout(content_layout)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

        # Apply stylesheet
        self.setStyleSheet("""
            QMainWindow {
                background-color: #1a1a2e;
            }
            QTableWidget {
                background-color: #16213e;
                color: #e6e6e6;
                gridline-color: #0f3460;
                border: 1px solid #0f3460;
                border-radius: 4px;
            }
            QHeaderView::section {
                background-color: #0f3460;
                color: white;
                padding: 5px;
                border: none;
                font-weight: bold;
            }
            QLabel {
                color: #f05454;
                font-family: 'Segoe UI';
            }
            QLineEdit {
                background-color: #16213e;
                color: white;
                padding: 8px;
                border: 1px solid #0f3460;
                border-radius: 4px;
                font-size: 14px;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #f05454;
                color: white;
            }
            QStatusBar {
                background-color: #0f3460;
                color: white;
                font-size: 12px;
            }
        """)

        self.all_packets = []  # Store all packet dicts
        signal_manager.new_packet.connect(self.add_packet)
        signal_manager.alert_signal.connect(self.show_alert)
        
        # Initialize counters
        self.packet_count = 0
        self.threat_count = 0

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.drag_pos = event.globalPos()
            event.accept()

    def mouseMoveEvent(self, event):
        if hasattr(self, 'drag_pos'):
            self.move(self.pos() + event.globalPos() - self.drag_pos)
            self.drag_pos = event.globalPos()
            event.accept()

    def mouseReleaseEvent(self, event):
        if hasattr(self, 'drag_pos'):
            del self.drag_pos

    def filter_table(self):
        text = self.filter_bar.text().strip().lower()
        protocols = []

        for row in range(self.packet_table.rowCount()):
            src_item = self.packet_table.item(row, 0)
            dst_item = self.packet_table.item(row, 1)
            proto_item = self.packet_table.item(row, 2)
            row_visible = False

            if src_item and dst_item:
                if text in src_item.text().lower() or text in dst_item.text().lower():
                    row_visible = True

            self.packet_table.setRowHidden(row, not row_visible)
            if row_visible and proto_item:
                protocols.append(proto_item.text())

        self.traffic_plot.update_protocol_counts(protocols)

    def animate_row_height(table, row, start=0, end=24, duration=300):
        steps = 10
        delay = duration // steps
        delta = (end - start) / steps

        def increase_height(step=0):
            if step <= steps:
                height = int(start + delta * step)
                table.setRowHeight(row, height)
                QTimer.singleShot(delay, lambda: increase_height(step + 1))

        increase_height()

    def add_packet(self, pkt):
        self.all_packets.append(pkt)
        self.packet_count += 1
        self.packet_count_label.setText(f"Packets: {self.packet_count}")

        row = self.packet_table.rowCount()
        self.packet_table.insertRow(row)
        self.packet_table.setItem(row, 0, QTableWidgetItem(pkt["src"]))
        self.packet_table.setItem(row, 1, QTableWidgetItem(pkt["dst"]))
        self.packet_table.setItem(row, 2, QTableWidgetItem(pkt["proto"]))
        self.packet_table.setItem(row, 3, QTableWidgetItem(pkt["port"]))
        self.packet_table.setItem(row, 4, QTableWidgetItem(pkt["flags"]))
        self.packet_table.setItem(row, 5, QTableWidgetItem(pkt["time"]))

        items = [
            QTableWidgetItem(pkt["src"]),
            QTableWidgetItem(pkt["dst"]),
            QTableWidgetItem(pkt["proto"]),
            QTableWidgetItem(pkt["port"]),
            QTableWidgetItem(pkt["flags"]),
            QTableWidgetItem(pkt["time"])
        ]

        if pkt["proto"] == "TCP":
            color = QColor(105, 10, 225)  # Royal Blue
        elif pkt["proto"] == "UDP":
            color = QColor(50, 205, 50)   # Lime Green
        else:
            color = QColor(169, 169, 169) # Gray
        if pkt["src"] in SUSPICIOUS_IPS or pkt["dst"] in SUSPICIOUS_IPS:
            color = QColor(220, 20, 60)  # Crimson
            self.threat_count += 1
            self.threat_count_label.setText(f"Threats: {self.threat_count}")
        
        for col, item in enumerate(items):
            item.setBackground(color.darker(150))
            item.setForeground(QColor(255, 255, 255))
            self.packet_table.setItem(row, col, item)
            
            # Add tooltips
            tooltip_text = f"""
            <b>Detailed Packet Info:</b><br>
            Source: {pkt['src']}<br>
            Destination: {pkt['dst']}<br>
            Protocol: {pkt['proto']}<br>
            Port: {pkt['port']}<br>
            Flags: {pkt['flags']}<br>
            Time: {pkt['time']}
            """
            item.setToolTip(tooltip_text)
        
        self.traffic_plot.increment()
        self.filter_table()  # Refresh protocol chart


    def show_alert(self, message):
        self.alert_label.setText(message)
        self.alert_label.setGraphicsEffect(QGraphicsOpacityEffect())
        self.alert_label.graphicsEffect().setOpacity(1.0)
        
        self.fade_animation = QPropertyAnimation(self.alert_label.graphicsEffect(), b"opacity")
        self.fade_animation.setDuration(10000)  # 10 seconds
        self.fade_animation.setStartValue(1.0)
        self.fade_animation.setEndValue(0.0)
        self.fade_animation.finished.connect(lambda: self.alert_label.setText(""))
        self.fade_animation.start()

def log_packet(pkt):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Logs"]
    ws.append([pkt["time"], pkt["src"], pkt["dst"], pkt["proto"], pkt["port"], pkt["flags"]])
    wb.save(EXCEL_FILE)

def analyze_packet(packet):
    if IP in packet:
        ip_layer = packet[IP]
        src = ip_layer.src
        dst = ip_layer.dst
        proto = "OTHER"
        port = "-"
        flags = "-"
        info = ""
        timestamp = datetime.now().strftime("%H:%M:%S")

        if TCP in packet:
            proto = "TCP"
            port = f"{packet[TCP].sport}->{packet[TCP].dport}"
            flags = str(packet[TCP].flags)

            if 'S' in flags:  # SYN scan
                connection_tracker.setdefault(src, []).append(packet[TCP].dport)

        elif UDP in packet:
            proto = "UDP"
            port = f"{packet[UDP].sport}->{packet[UDP].dport}"

        pkt = {
            "src": src,
            "dst": dst,
            "proto": proto,
            "port": port,
            "flags": flags,
            "time": timestamp
        }

        signal_manager.new_packet.emit(pkt)
        log_packet(pkt)

        if src in SUSPICIOUS_IPS:
            signal_manager.alert_signal.emit(f"[!] Suspicious IP Detected: {src}")

def detect_port_scans():
    while True:
        for ip, ports in list(connection_tracker.items()):
            if len(set(ports)) > 10:
                signal_manager.alert_signal.emit(f"[!] Port Scan Detected from {ip} on ports {sorted(set(ports))}")
                del connection_tracker[ip]
        time.sleep(10)

def start_sniffing():
    sniff(prn=analyze_packet, filter="ip", store=False)

def main():
    app = QApplication(sys.argv)
    gui = NetworkAnalyzerGUI()
    gui.show()

    threading.Thread(target=start_sniffing, daemon=True).start()
    threading.Thread(target=detect_port_scans, daemon=True).start()

    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
